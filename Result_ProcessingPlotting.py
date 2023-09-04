from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.chart import ScatterChart, Reference, series


def DuplicateSheet(originalSheet, newSheet):
    maxrow = originalSheet.max_row
    max_col = originalSheet.max_column

    for row in range(1, maxrow + 1):
        for col in range(1, max_col + 1):
            newSheet.cell(row=row, column=col, value=originalSheet.cell(row=row, column=col).value)

def open_folder_dialog():
    root = tk.Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(
        title="Select the root folder containing the results"
    )

    return folder_path

def list_subfolders(root_folder):
    subfolders = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        for dirname in dirnames:
            subfolder_path = os.path.join(dirpath, dirname)
            subfolders.append(subfolder_path)

    return subfolders

def filePaths(Subfolders, FileName):
    ResultFiles = []
    for subfolder in subfolders:
        for files in os.listdir(subfolder):
            if files.endswith(FileName):
                Filepath = os.path.join(subfolder, files)
                ResultFiles.append(Filepath)

    return ResultFiles

def ExcelObj(OutputPath, Sheetname):
    # Create a new Excel workbook
    if os.path.exists(OutputPath):
        workbook = openpyxl.load_workbook(OutputPath)

        if Sheetname in workbook.sheetnames:
            worksheet = workbook[Sheetname]
        else:
            worksheet = workbook.create_sheet(title=Sheetname)

        return workbook, worksheet

    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title=Sheetname)
        return workbook, worksheet


def ExcelWriter(ResultFiles, OutputPath, Sheetname):
    StartingRow = 6

    LiveRow = StartingRow

    workbook, worksheet = ExcelObj(OutputPath, Sheetname )



    for file in ResultFiles:
        SubjectCode = file.split(os.path.sep)[-4:-1]

        fObj = open(file, "r")

        lines = fObj.read()

        # Split the content into rows and cells based on tabs and newlines
        rows = lines.split('\n')
        cells_in_rows = [row.split('\t') for row in rows]

        # Convert the cells to float values
        float_cells = []
        for row in cells_in_rows:
            float_row = []
            for cell_value in row:
                try:
                    float_value = float(cell_value)
                except ValueError:
                    float_value = cell_value
                float_row.append(float_value)
            float_cells.append(float_row)





        # Write the cells into the Excel sheet
        #Writing for the subject codes
        for index, value in enumerate(SubjectCode):
            worksheet.cell(row=LiveRow, column=index + 1, value=value)
        LiveRow += 1


        #Writing for the content
        for row_index, row in enumerate(float_cells, start=LiveRow):
            for col_index, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=col_index, value=cell_value)
            LiveRow += 1
        LiveRow += 2

    workbook.save(OutputPath)

def List_Intersection(list1, list2):
    intersection = list(set(list1).intersection(list2))
    intersection.sort()

    return intersection

def TableRowsFinder(ws):
        wsTableRows = []
        maxRow = ws.max_row
        startTable = False
        for row in range(1, 1+ maxRow):
            value = ws.cell(row = row, column = 1).value
            if value != None and startTable:
                startTable = True


            elif value != None:
                startTable = True

                wsTableRows.append(row)
            elif value == None:
                startTable = False
        return  wsTableRows

def rowContains_List(worksheet, Keyword):
    Rows = []
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column


    for row in range(1, max_rows+1):
        for column in range(1, max_cols):
            Value = worksheet.cell(row=row, column= column).value
            if Keyword == Value:
                Rows.append(row)
    return Rows

def colContains_List(worksheet, Keyword):
    Cols = []
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column


    for row in range(1, max_rows+1):
        for column in range(1, max_cols):
            Value = worksheet.cell(row=row, column= column).value
            if Keyword == Value:
                Cols.append(column)


    return list(set(Cols))
def tableExtent(worksheet, row):
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column
    RowStart = row

    # Column Extent
    Column = 0
    for row in range(RowStart, max_rows+1):
        rowValue = None
        for column in range(1, max_cols+10):
            Value = worksheet.cell(row=row, column= column).value

            if Value == None or Value == "None":
                if Column < column:
                    Column  = column
                break

            else:
                rowValue = "NotNone"


        if rowValue == None:
            break

    # Row Extent
    Row = 0
    for column in range(1, max_cols + 1):
        colValue = None
        for row in range(RowStart, max_rows+10):
            Value = worksheet.cell(row=row, column= column).value
            if Value is None:
                if Row < column:
                    Row  = row
                break
            else:
                colValue = "NotNone"
        if colValue == None:
            break

    RowEnd = Row
    ColumnStart = 1
    ColumnEnd = Column


    return RowStart, RowEnd, ColumnStart, ColumnEnd

def GroupWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols, skip_TitleRows, CustomCol = [0]):
    #CustomCol = [0] indicates no need for adding Customcolumn (If custom column is provided never provide initial item as 0
    # Writing Content
    # Code Writer
    for index, code in enumerate(SubjectCode):
        ws.cell(row=LiveRow, column=index + 1, value=code)
    LiveRow += 1

    # Identifier Writer
    if CustomCol[0] == 0:
        for index, col in enumerate(range(1, len(SelectedRows) + 1)):
            ws.cell(row=LiveRow, column=index + 1,
                    value=worksheet.cell(row=SelectedRows[index][0], column=IdentifierCol).value)
        LiveRow += 1

    else:
        for index, col in enumerate(range(1, len(SelectedRows) + 2)):
            if index == 0:
                ws.cell(row=LiveRow, column=index + 1,
                        value=CustomCol[index])

            else:
                ws.cell(row=LiveRow, column=index + 1,
                        value=worksheet.cell(row=SelectedRows[index - 1][0], column=IdentifierCol).value)
        LiveRow += 1


    # Content Writer
    XValues = []

    if CustomCol[0] == 0:
        for sourCol in SelectedCols:
            for Cindex, col in enumerate(range(1, len(SelectedRows) + 1)):
                Xvalue = []
                for Rindex, row in enumerate(
                        range(SelectedRows[Cindex][0] + skip_TitleRows, SelectedRows[Cindex][1])):
                    value = worksheet.cell(row=row, column=sourCol).value
                    ws.cell(row=LiveRow + Rindex, column=Cindex + 1, value=value)

                    Xvalue.append(value)

                XValues.append(Xvalue)



            LiveRow = LiveRow + (Rindex + 1) + 2

    else:
        for sourCol in SelectedCols:
            for Cindex, col in enumerate(range(1, len(SelectedRows) + 2)):
                if Cindex == 0:
                    for index, row in enumerate(range(len(CustomCol) - 1) ):
                            ws.cell(row=LiveRow + index, column=Cindex + 1, value=CustomCol[index + 1])


                else:
                    Xvalue = []
                    for Rindex, row in enumerate(
                            range(SelectedRows[Cindex-1][0] + skip_TitleRows, SelectedRows[Cindex-1][1])):
                        print(sourCol)
                        value = float(worksheet.cell(row=row, column=sourCol).value)
                        print(value, sourCol)
                        ws.cell(row=LiveRow + Rindex, column=Cindex + 1,  value=value)

                        Xvalue.append(value)
                    XValues.append(Xvalue)

            LiveRow = LiveRow + (Rindex + 1) + 7

    return LiveRow, XValues

def Charts(ws, YValues, XValues, XLabel, YLabel,SeriesLabel, AnchourCell = "",  Title = "" ):
    items = 0
    properties = 0

    chart = openpyxl.chart.ScatterChart()
    chart.title = Title
    chart.style = 4
    chart.x_axis.title = XLabel
    chart.y_axis.title = YLabel

    xvalues = None
    values = None
    series = None
    print(YValues)


    if len(YValues[0]) == 1:
        y_values = Reference(ws, min_col=YValues[0][0], min_row=YValues[1][0], max_row=YValues[1][1])

        for index, column in enumerate(XValues[0]):
            x_values = Reference(ws, min_col=column, min_row=XValues[1][0], max_row=XValues[1][1])
            series = openpyxl.chart.Series(y_values, x_values, title=SeriesLabel[index])
            chart.series.append(series)


            #Markers and Curves
            series = chart.series[-1]
            Marker = Markers[index]
            series.marker.symbol = Marker
            series.marker.size = 7
            # series.SeriesLabel = SeriesLabel[index]

            if Marker == "star" or Marker == "plus" or Marker == 'x':
                series.marker.graphicalProperties.line.solidFill = SeriesColour[index]
                series.marker.graphicalProperties.line.width = pixels_to_EMU(2.5)
            else:
                series.marker.graphicalProperties.solidFill = SeriesColour[index]
                series.marker.graphicalProperties.line.noFill = True

            series.graphicalProperties.line.solidFill = SeriesColour[index]
            series.graphicalProperties.line.dashStyle = Dashes[0]
            series.graphicalProperties.line.width = pixels_to_EMU(2.5)

            chart.legend.position = 'b'


    #Axis and Paragraph texts
    pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))])
    chart.x_axis.txPr = rtp
    chart.y_axis.txPr = rtp
    pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))])
    chart.legend.txPr = rtp
    chart.title.tx.rich.p[0].pPr = pp
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp
    # Adjust graph size
    chart.width = 12     #Elselvier page halfwidth = 9cm        Writing area Only
    chart.height = 8   #Elselvier page full height = 24 cm        Writing area Only






    if AnchourCell == "":
        ws.add_chart(chart, "H1")

    else:
        ws.add_chart(chart, AnchourCell)

def DataSorting(refSheet, workingSheet, sortDatabase, sortRef , alsoSortfor, SelectedTable_Rows,  skip_TitleRows, sort_Ascending = False):
    worksheet = refSheet
    ws = workingSheet

    Identifiers = sortDatabase

    # Preferred to use single Selected Cols for single sheet i.e Group_A
    SelectedCols = colContains_List(worksheet, sortRef)
    IdentifierCol = colContains_List(worksheet, Identifiers[0])[0]

    # Selected rows list (Contains the value with rowStart and rowEnd)
    SelectedRows = []
    for row in SelectedTable_Rows:
        print(row, SelectedTable_Rows)
        RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
        SelectedRows.append([RowStart, RowEnd])

    # Getting the values of the indicators in the List [List_ind1, List_ind2,....] ie of the ReferenceSort data
    valueIndicators = []
    for rows in SelectedRows:
        startRow = rows[0]
        endRow = rows[1]
        valueList = []
        for row in range(startRow + skip_TitleRows, endRow):
            for col in SelectedCols:
                cellvalue = worksheet.cell(row=row, column=col).value
                valueList.append(cellvalue)
        valueIndicators.append(valueList)

    refValue = []
    for values in valueIndicators:
        refValue.append(values[0])

    # Sorting in descending order and writing for the corresponding index
    dynRefValue = [x for x in refValue]
    if sort_Ascending:
        dynRefValue.sort()

    else:
        dynRefValue.sort(reverse=True)

    # New Index of the sorted data is placed to the previous value position
    sortedIndex = []
    for value in dynRefValue:
        sortedIndex.append(refValue.index(value))

    # Placing Table row - Used for placing sequentially to the sorted database
    SelectedTable_Rows = [x[0] for x in SelectedRows]
    placingRows = []
    for data in sortDatabase:
        for row in SelectedTable_Rows:
            rowValue = worksheet.cell(column=IdentifierCol, row=row).value
            if rowValue == data:
                RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
                placingRows.append([RowStart, RowEnd])

    # Writing to the Placing Table row in counter of the sorted data both being in seqential manner
    for origInd, ind in enumerate(sortedIndex):
        rowStart = placingRows[origInd][0]
        rowEnd = placingRows[origInd][1]
        for col in SelectedCols:
            for index, row in enumerate(range(rowStart + skip_TitleRows, rowEnd)):
                ws.cell(row=row, column=col).value = valueIndicators[ind][index]

    # Writing for the alsoSortFor dataset based completely on ref(Tailed to the reference dataset)
    alsoSortCols = []
    for alsoSort in alsoSortfor:
        alsoSortCols.append(colContains_List(worksheet, alsoSort)[0])
    print(alsoSortCols)

    for origInd, ind in enumerate(sortedIndex):
        rowStart = SelectedRows[ind][0]
        rowEnd = SelectedRows[ind][1]
        for col in alsoSortCols:
            for index, row in enumerate(range(rowStart + skip_TitleRows, rowEnd)):
                ws.cell(row= placingRows[origInd][0] + index + skip_TitleRows, column=col).value = worksheet.cell(row=row, column=col).value


def Assembly_Correction(OutputPath, Sheetname, Keyword = "Reaction Force X"):
    workbook, worksheet = ExcelObj(OutputPath, Sheetname)


    sortRef = Keyword
    alsoSortfor = ["Reaction Force Y","Reaction Moment X", "Reaction Moment Y"]        #The data will also be tailed based on sorted database and its corresponding index

    Identifiers = Soils
    skip_TitleRows = 2
    #Sort database are also known as identifiers in the first row of table
    # The sorted data will be placed in descending order for provided list
    # if the sortDatabase =["Fixed", "Hard", ___], and sort ascending is true, then the lowest value is placed to the fixed type of table and so on
    sortDatabase = ["Fixed", "Hard", "Medium", "Soft"]

    GroupSheet = f"{Sheetname}_Refined"
    workbook, ws = ExcelObj(OutputPath, GroupSheet)
    ws.cell(row=2, column=2, value=Keyword)
    wsTableRows = []

    #Duplicate the correction slide
    DuplicateSheet(originalSheet = worksheet, newSheet = ws)


    #Writing
    for index, building in enumerate(Buildings):
        for earthquake in Earthquakes:
            SubjectCode = [building, earthquake]

            BuildingRows = rowContains_List(worksheet, building)
            EarthquakeRows = rowContains_List(worksheet, earthquake)
            SelectedTable_Rows = List_Intersection(BuildingRows, EarthquakeRows)

            DataSorting(worksheet, ws, sortDatabase, sortRef, alsoSortfor, SelectedTable_Rows, skip_TitleRows, sort_Ascending = False)



    workbook.save(OutputPath)


#Same Building, Same Earthquake, Different Soil (Soil is Identifier)
def Group_A(Keyword = "Drift X"):
    def Correction(ws, ValueList, CorrectionRow, CorrType, SurrRefRow=1):
        for col in ValueList[0]:
            print(ValueList, CorrectionRow)
            IntRow = ValueList[1][0] + CorrectionRow - 1
            Ref1 = ws.cell(row=IntRow - 1, column=col).value
            Ref2 = ws.cell(row=IntRow + 1, column=col).value

            if CorrType == "Average":
                IntRowValue = Ref2 + 0.7 * (Ref1 - Ref2)

            ws.cell(row=IntRow, column=col, value=IntRowValue)
    workbook, worksheet = ExcelObj(OutputPath, Sheetname)

    CustomColumn = ["Building height (m)",0, 4,7,10,13,16]
    LiveRow = 6
    skip_TitleRows = 2
    Identifiers = Soils

    GroupSheet = Keyword
    workbook, ws = ExcelObj(OutputPath, GroupSheet)
    ws.cell(row=2, column=2, value=Keyword)
    wsTableRows = []


    #Writing
    for building in Buildings:
        for earthquake in Earthquakes:
            SubjectCode = [building, earthquake]

            BuildingRows = rowContains_List(worksheet, building)
            EarthquakeRows = rowContains_List(worksheet, earthquake)

            SelectedTable_Rows= List_Intersection(BuildingRows, EarthquakeRows)

            # Preferred to use single Selected Cols for single sheet i.e Group_A
            SelectedCols = colContains_List(worksheet, Keyword)
            IdentifierCol = colContains_List(worksheet, Identifiers[0])[0]


            #Selected rows list (Contains the value with rowStart and rowEnd)
            SelectedRows = []
            for row in SelectedTable_Rows:
                RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
                SelectedRows.append([RowStart, RowEnd])



            #Writer
            CurrentTableRow = LiveRow
            wsTableRows.append(CurrentTableRow)
            LiveRow, XValues = GroupWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols, skip_TitleRows, CustomCol = CustomColumn)




    #Make sure to save file after calling GroupWriter
    workbook.save(OutputPath)

    # Table Info
    TableInfo = []
    for row in wsTableRows:
        RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(ws, row)
        TableInfo.append([RowStart, RowEnd, ColumnStart, ColumnEnd])

    # Correction
    if Keyword.split(" ")[0] == "Drift":
        CorrectionRow = 4
        SurrRefRow = 1
        for info in TableInfo:
            RowStart, RowEnd, ColumnStart, ColumnEnd = info[0], info[1], info[2], info[3]

            ValuesList = [[], [RowStart + skip_TitleRows, RowEnd]]
            for i in range(ColumnStart + 1, ColumnEnd):
                ValuesList[0].append(i)
            CorrType = "Average"


            Correction(ws, ValuesList, CorrectionRow,CorrType,  SurrRefRow=SurrRefRow)



    # Plotter
    for info in TableInfo:
        RowStart, RowEnd, ColumnStart, ColumnEnd = info[0], info[1], info[2], info[3]
        # YValues =  [Column ][Row Extents]
        YValues = [[ColumnStart], [RowStart + skip_TitleRows, RowEnd]]
        XValues = [[], [RowStart + skip_TitleRows, RowEnd]]
        for i in range(ColumnStart + 1, ColumnEnd):
            XValues[0].append(i)

        Anchor = f"{get_column_letter(ColumnEnd + 2)}{RowStart}"

        key = Keyword.split()[0]
        direction = Keyword.split()[1]
        BuildingName = ws.cell(row=RowStart, column=1).value
        EQName = ws.cell(row=RowStart, column=2).value
        CTitle  = f"{key} of {BuildingName} building for {EQName} earthquake along {direction}"

        if Keyword.split()[0] == "Displacement":
            XLabel = Keyword.split()[0] + " (m)"

        elif Keyword.split()[0] == "Drift":
            XLabel = Keyword.split()[0]
        else:
            XLabel = Keyword

        Charts(ws, YValues, XValues, XLabel=XLabel, YLabel=CustomColumn[0],SeriesLabel = Soils, AnchourCell=Anchor, Title=CTitle)

    workbook.save(OutputPath)


def Group_B(Keyword="Reaction Force X"):
    # Same Earthquake,Different Soils, Different Building__(Point data for each soil ie Max)
    # Generally used for base shear, base moment
    workbook, worksheet = ExcelObj(OutputPath, Sheetname)

    CustomColumn = ["Building height (m)", 0, 4, 7, 10, 13, 16]
    LiveRow = 6
    skip_TitleRows = 2
    PointRow = 0  # 0 Indicates for the first value row
    Identifiers = [Soils, Buildings]

    GroupSheet = Keyword
    workbook, ws = ExcelObj(OutputPath, GroupSheet)
    ws.cell(row=2, column=2, value=Keyword)
    wsTableRows = []

    # Writing
    for earthquake in Earthquakes:
        TitleIteration = True
        for building in Buildings:

            SubjectCode = [earthquake]
            BuildingRows = rowContains_List(worksheet, building)
            EarthquakesRows = rowContains_List(worksheet, earthquake)

            # BuildingRows = rowContains_List(worksheet, building)
            SelectedTable_Rows = List_Intersection(BuildingRows, EarthquakesRows)
            # SelectedTable_Rows= List_Intersection(BuildingRows, EarthquakeRows)
            # Preferred to use single Selected Cols for single sheet i.e. Group_A
            SelectedCols = colContains_List(worksheet, Keyword)

            IdentifierCol = []

            for identifier in Identifiers:
                IdentifierCol.append(colContains_List(worksheet, identifier[0])[0])

            # IdentifierCol = colContains_List(worksheet, Identifiers[0])[0]
            # Selected rows list (Contains the value with rowStart and rowEnd)
            SelectedRows = []
            for row in SelectedTable_Rows:
                RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
                SelectedRows.append([RowStart, RowEnd])

            # Writer
            if TitleIteration:
                CurrentTableRow = LiveRow
                wsTableRows.append(CurrentTableRow)

            def PointWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols,
                            skip_TitleRows):
                # CustomCol = [0] indicates no need for adding Customcolumn (If custom column is provided never provide initial item as 0
                # Writing Content
                IdentifierColY = colContains_List(worksheet, Buildings[0])[0]
                CustomColumn = ["Buildings"]

                # Code Writer
                if TitleIteration:
                    for index, code in enumerate(SubjectCode):
                        ws.cell(row=LiveRow, column=index + 1, value=code)
                    LiveRow += 1

                    # Identifier Writer
                    for index, col in enumerate(range(1, len(SelectedRows) + 2)):
                        if index == 0:
                            ws.cell(row=LiveRow, column=index + 1,
                                    value=CustomColumn[index])
                        else:
                            ws.cell(row=LiveRow, column=index + 1,
                                    value=worksheet.cell(row=SelectedRows[index - 1][0], column=IdentifierCol[0]).value)
                    LiveRow += 1

                # # Content Writer
                CustomColumn = worksheet.cell(row=SelectedRows[0][0], column=IdentifierColY).value

                XValues = []

                for sourCol in SelectedCols:
                    for Cindex, col in enumerate(range(1, len(SelectedRows) + 2)):
                        if Cindex == 0:
                            ws.cell(row=LiveRow, column=Cindex + 1, value=CustomColumn)
                        else:

                            value = float(
                                worksheet.cell(row=SelectedRows[Cindex - 1][0] + skip_TitleRows, column=sourCol).value)
                            ws.cell(row=LiveRow, column=Cindex + 1, value=value)
                            # Xvalue.append(value)
                            # XValues.append(Xvalue)
                    LiveRow = LiveRow + 1

                return LiveRow

            LiveRow = PointWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols,
                                  skip_TitleRows)
            TitleIteration = False

        if TitleIteration == False:
            LiveRow += 1

    workbook.save(OutputPath)

    #
    # #Make sure to save file after calling GroupWriter
    # workbook.save(OutputPath)
    #
    # # Table Info
    TableInfo = []
    for row in wsTableRows:
        RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(ws, row)
        TableInfo.append([RowStart, RowEnd, ColumnStart, ColumnEnd])

    # # Plotter
    for info in TableInfo:
        RowStart, RowEnd, ColumnStart, ColumnEnd = info[0], info[1], info[2], info[3]
        # YValues =  [Column ][Row Extents]
        YValues = [[ColumnStart], [RowStart + skip_TitleRows, RowEnd]]
        XValues = [[], [RowStart + skip_TitleRows, RowEnd]]
        for i in range(ColumnStart + 1, ColumnEnd):
            XValues[0].append(i)

        Anchor = f"{get_column_letter(ColumnEnd + 2)}{RowStart}"

        key = Keyword.split()[0]
        direction = Keyword.split()[1]
        BuildingName = ws.cell(row=RowStart, column=1).value
        EQName = ws.cell(row=RowStart, column=2).value
        CTitle = f"{key} of {BuildingName} building for {EQName} earthquake along {direction}"

        if Keyword.split()[0] == "Displacement":
            XLabel = Keyword.split()[0] + " (m)"

        elif Keyword.split()[0] == "Drift":
            XLabel = Keyword.split()[0]
        else:
            XLabel = Keyword

        Charts(ws, YValues, XValues, XLabel=XLabel, YLabel=CustomColumn[0], SeriesLabel=Soils, AnchourCell=Anchor,
               Title=CTitle)

    workbook.save(OutputPath)





folderpath = open_folder_dialog()
subfolders = list_subfolders(folderpath)

FileName = "Result.txt"
ResultFiles = filePaths(subfolders, FileName)


OutputName =  "ResultAssembly"
Sheetname = "ResultAssembly"
OutputPath = os.path.join(folderpath, f'{OutputName}.xlsx')
ExcelWriter(ResultFiles, OutputPath, Sheetname)


Buildings = [ "S",  "L1", "L2", "L3", "L4", "R"]
Earthquakes = ["Gorkha", "Northridge", "San Fernando"]
Soils = ["Fixed", "Hard", "Medium",  "Soft"]
#ResultsKey = ["Drift X",'Drift Y','Displacement X','Displacement Y','Reaction Force X','Reaction Force Y','Reaction Moment X', 'Reaction Moment Y','Rocking Angle X','Rocking Angle Y','Rocking Angle Z','Max Uplift','Max Uplift Point','Max Settlement','Max Settlement Point']
ResultsKey = ["Drift X",'Drift Y','Displacement X','Displacement Y','Rotation X','Rotation Y','Rotation Z']

#Customized Arranger Chart List
SeriesColour = ["D30000", "0018F9", "3BB143", "FCE205", "FC0FC0", "2B1700", "000080", "FF00FF", "4CBB17", "C21807", "0080FE"]
Markers = [ "x", "triangle", "square",  "circle",  "plus", "diamond", "star", "dash", "dot",  "picture",  "auto"]
Dashes = ["solid", "sysDash", "dash", "dot", "sysDot", "lgDashDot",  "dashDot", "lgDashDotDot", "sysDashDotDot", "lgDash", "sysDashDot"]

# #
for key in ResultsKey:
    Group_A(Keyword = key)


# Assembly_Correction(OutputPath, Sheetname, Keyword = "Reaction Force X")
#




Group_B(Keyword = "Reaction Moment X")


#Normalizing of the values in the table
def Group_C(Keyword = "Reaction Force X"):
    # Same Earthquake,Different Soils, Different Building__(Point data for each soil ie Max)
    # Generally used for base shear, base moment
    workbook, worksheet = ExcelObj(OutputPath, Keyword)

    LiveRow = 6
    skip_TitleRows = 2
    PointRow = 0  # 0 Indicates for the first value row

    GroupSheet = f"Normalized {Keyword}"
    workbook, ws = ExcelObj(OutputPath, GroupSheet)
    ws.cell(row=2, column=2, value=Keyword)

    databook, datasheet = ExcelObj(os.path.join(folderpath, "Building Informations.xlsx"), "General Informations")



    wsTableRows = []
    wsTableRows = TableRowsFinder(worksheet)


    TableExtents = []
    for row in wsTableRows:
        RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
        TableExtents.append([RowStart, RowEnd, ColumnStart, ColumnEnd])

    IdentiferColumn = 1
    DuplicateSheet(worksheet, ws)

    def MultiplierComp(Identifer, IdentifierList,  corrDataList, Keyword = "Reaction Force X"):
        corrValue = corrDataList[IdentifierList.index(Identifer)]

        if Keyword.split(" ")[1] == "Force":
            Multiplier = 100/corrValue

        if Keyword.split(" ")[1] == "Moment":
            Multiplier = 100/(corrValue * 16)

        return Multiplier

    IdentifierList = []
    corrDataList = []


    #Data extraction for use in correction
    for index, col in enumerate(range(2, 4)):
        for row in range(3, 9):
            value = datasheet.cell(row = row, column = col).value

            if index == 0:
                IdentifierList.append(value)
            else:
                corrDataList.append(value)






    for celldata in TableExtents:
        RowStart, RowEnd, ColumnStart, ColumnEnd = celldata[0], celldata[1], celldata[2], celldata[3]
        Identifiers  = []

        for col in range(ColumnStart, ColumnEnd):
            for Rindex, row in enumerate(range(RowStart + skip_TitleRows, RowEnd)):

                if IdentiferColumn == col:
                        dynIdentifier = worksheet.cell(row=  row, column = IdentiferColumn).value
                        Identifiers.append(dynIdentifier)

                else:
                    cellvalue = worksheet.cell(row = row, column = col).value
                    print(cellvalue)

                    Identifer = Identifiers[Rindex]
                    Multiplier = MultiplierComp(Identifer, IdentifierList,  corrDataList, Keyword =Keyword)

                    resultVal = cellvalue * Multiplier
                    ws.cell(row=row, column=col, value = resultVal)


    workbook.save(OutputPath)




Group_C(Keyword = "Reaction Moment X")



