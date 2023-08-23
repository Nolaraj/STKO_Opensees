from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.chart import ScatterChart, Reference, series

def open_folder_dialog():
    root = tk.Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(
        title="Select the root folder containing the results"
    )

    return folder_path

def list_subfolders(root_folder):
    subfolders = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        for dirname in dirnames:
            subfolder_path = os.path.join(dirpath, dirname)
            subfolders.append(subfolder_path)

    return subfolders

def filePaths(Subfolders, FileName):
    ResultFiles = []
    for subfolder in subfolders:
        for files in os.listdir(subfolder):
            if files.endswith(FileName):
                Filepath = os.path.join(subfolder, files)
                ResultFiles.append(Filepath)

    return ResultFiles

def ExcelObj(OutputPath, Sheetname):
    # Create a new Excel workbook
    if os.path.exists(OutputPath):
        workbook = openpyxl.load_workbook(OutputPath)

        if Sheetname in workbook.sheetnames:
            worksheet = workbook[Sheetname]
        else:
            worksheet = workbook.create_sheet(title=Sheetname)

        return workbook, worksheet

    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title=Sheetname)
        return workbook, worksheet


def ExcelWriter(ResultFiles, OutputPath, Sheetname):
    StartingRow = 6

    LiveRow = StartingRow

    workbook, worksheet = ExcelObj(OutputPath, Sheetname )



    for file in ResultFiles:
        SubjectCode = file.split(os.path.sep)[-4:-1]

        fObj = open(file, "r")

        lines = fObj.read()

        # Split the content into rows and cells based on tabs and newlines
        rows = lines.split('\n')
        cells_in_rows = [row.split('\t') for row in rows]

        # Convert the cells to float values
        float_cells = []
        for row in cells_in_rows:
            float_row = []
            for cell_value in row:
                try:
                    float_value = float(cell_value)
                except ValueError:
                    float_value = cell_value
                float_row.append(float_value)
            float_cells.append(float_row)





        # Write the cells into the Excel sheet
        #Writing for the subject codes
        for index, value in enumerate(SubjectCode):
            worksheet.cell(row=LiveRow, column=index + 1, value=value)
        LiveRow += 1


        #Writing for the content
        for row_index, row in enumerate(float_cells, start=LiveRow):
            for col_index, cell_value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=col_index, value=cell_value)
            LiveRow += 1
        LiveRow += 2

    workbook.save(OutputPath)

def List_Intersection(list1, list2):
    intersection = list(set(list1).intersection(list2))
    intersection.sort()

    return intersection

def rowContains_List(worksheet, Keyword):
    Rows = []
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column


    for row in range(1, max_rows+1):
        for column in range(1, max_cols):
            Value = worksheet.cell(row=row, column= column).value
            if Keyword == Value:
                Rows.append(row)
    return Rows

def colContains_List(worksheet, Keyword):
    Cols = []
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column


    for row in range(1, max_rows+1):
        for column in range(1, max_cols):
            Value = worksheet.cell(row=row, column= column).value
            if Keyword == Value:
                Cols.append(column)


    return list(set(Cols))
def tableExtent(worksheet, row):
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column
    RowStart = row

    # Column Extent
    Column = 0
    for row in range(RowStart, max_rows+1):
        rowValue = None
        for column in range(1, max_cols+10):
            Value = worksheet.cell(row=row, column= column).value

            if Value == None or Value == "None":
                if Column < column:
                    Column  = column
                break

            else:
                rowValue = "NotNone"


        if rowValue == None:
            break

    # Row Extent
    Row = 0
    for column in range(1, max_cols + 1):
        colValue = None
        for row in range(RowStart, max_rows+10):
            Value = worksheet.cell(row=row, column= column).value
            if Value is None:
                if Row < column:
                    Row  = row
                break
            else:
                colValue = "NotNone"
        if colValue == None:
            break

    RowEnd = Row
    ColumnStart = 1
    ColumnEnd = Column


    return RowStart, RowEnd, ColumnStart, ColumnEnd

def GroupWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols, skip_TitleRows, CustomCol = [0]):
    #CustomCol = [0] indicates no need for adding Customcolumn (If custom column is provided never provide initial item as 0
    # Writing Content
    # Code Writer
    for index, code in enumerate(SubjectCode):
        ws.cell(row=LiveRow, column=index + 1, value=code)
    LiveRow += 1

    # Identifier Writer
    if CustomCol[0] == 0:
        for index, col in enumerate(range(1, len(SelectedRows) + 1)):
            ws.cell(row=LiveRow, column=index + 1,
                    value=worksheet.cell(row=SelectedRows[index][0], column=IdentifierCol).value)
        LiveRow += 1

    else:
        for index, col in enumerate(range(1, len(SelectedRows) + 2)):
            if index == 0:
                ws.cell(row=LiveRow, column=index + 1,
                        value=CustomCol[index])

            else:
                ws.cell(row=LiveRow, column=index + 1,
                        value=worksheet.cell(row=SelectedRows[index - 1][0], column=IdentifierCol).value)
        LiveRow += 1


    # Content Writer
    XValues = []

    if CustomCol[0] == 0:
        for sourCol in SelectedCols:
            for Cindex, col in enumerate(range(1, len(SelectedRows) + 1)):
                Xvalue = []
                for Rindex, row in enumerate(
                        range(SelectedRows[Cindex][0] + skip_TitleRows, SelectedRows[Cindex][1])):
                    value = worksheet.cell(row=row, column=sourCol).value
                    ws.cell(row=LiveRow + Rindex, column=Cindex + 1, value=value)

                    Xvalue.append(value)

                XValues.append(Xvalue)



            LiveRow = LiveRow + (Rindex + 1) + 2

    else:
        for sourCol in SelectedCols:
            for Cindex, col in enumerate(range(1, len(SelectedRows) + 2)):
                if Cindex == 0:
                    for index, row in enumerate(range(len(CustomCol) - 1) ):
                            ws.cell(row=LiveRow + index, column=Cindex + 1, value=CustomCol[index + 1])


                else:
                    Xvalue = []
                    for Rindex, row in enumerate(
                            range(SelectedRows[Cindex-1][0] + skip_TitleRows, SelectedRows[Cindex-1][1])):
                        value = float(worksheet.cell(row=row, column=sourCol).value)
                        ws.cell(row=LiveRow + Rindex, column=Cindex + 1,  value=value)

                        Xvalue.append(value)
                    XValues.append(Xvalue)

            LiveRow = LiveRow + (Rindex + 1) + 7

    return LiveRow, XValues

def Charts(ws, YValues, XValues, XLabel, YLabel,SeriesLabel, AnchourCell = "",  Title = "" ):
    items = 0
    properties = 0

    chart = openpyxl.chart.ScatterChart()
    chart.title = Title
    chart.style = 4
    chart.x_axis.title = XLabel
    chart.y_axis.title = YLabel

    xvalues = None
    values = None
    series = None
    print(YValues)


    if len(YValues[0]) == 1:
        y_values = Reference(ws, min_col=YValues[0][0], min_row=YValues[1][0], max_row=YValues[1][1])

        for index, column in enumerate(XValues[0]):
            x_values = Reference(ws, min_col=column, min_row=XValues[1][0], max_row=XValues[1][1])
            series = openpyxl.chart.Series(y_values, x_values, title=SeriesLabel[index])
            chart.series.append(series)


            #Markers and Curves
            series = chart.series[-1]
            Marker = Markers[index]
            series.marker.symbol = Marker
            series.marker.size = 7
            # series.SeriesLabel = SeriesLabel[index]

            if Marker == "star" or Marker == "plus" or Marker == 'x':
                series.marker.graphicalProperties.line.solidFill = SeriesColour[index]
                series.marker.graphicalProperties.line.width = pixels_to_EMU(2.5)
            else:
                series.marker.graphicalProperties.solidFill = SeriesColour[index]
                series.marker.graphicalProperties.line.noFill = True

            series.graphicalProperties.line.solidFill = SeriesColour[index]
            series.graphicalProperties.line.dashStyle = Dashes[0]
            series.graphicalProperties.line.width = pixels_to_EMU(2.5)

            chart.legend.position = 'b'


    #Axis and Paragraph texts
    pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=900, b=False))])
    chart.x_axis.txPr = rtp
    chart.y_axis.txPr = rtp
    pp = ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=CharacterProperties(latin=Font(typeface='Times New Roman'), sz=1000, b=True))])
    chart.legend.txPr = rtp
    chart.title.tx.rich.p[0].pPr = pp
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp
    # Adjust graph size
    chart.width = 12     #Elselvier page halfwidth = 9cm        Writing area Only
    chart.height = 8   #Elselvier page full height = 24 cm        Writing area Only






    if AnchourCell == "":
        ws.add_chart(chart, "H1")

    else:
        ws.add_chart(chart, AnchourCell)


def Correction(ws, ValueList,CorrectionRow, CorrType, SurrRefRow = 1):
        for col in ValueList[0]:
            print(ValueList,CorrectionRow)
            IntRow = ValueList[1][0] + CorrectionRow - 1
            Ref1 = ws.cell(row=IntRow - 1, column=col).value
            Ref2 = ws.cell(row=IntRow + 1, column=col).value

            if CorrType == "Average":
                IntRowValue = Ref2 + 0.7*(Ref1 - Ref2)

            ws.cell(row=IntRow , column=col, value = IntRowValue)


#Same Building, Same Earthquake, Different Soil (Soil is Identifier)
def Group_A(Keyword = "Drift X"):
    workbook, worksheet = ExcelObj(OutputPath, Sheetname)

    CustomColumn = ["Building height (m)",0, 4,7,10,13,16]
    LiveRow = 6
    skip_TitleRows = 2
    Identifiers = Soils

    GroupSheet = Keyword
    workbook, ws = ExcelObj(OutputPath, GroupSheet)
    ws.cell(row=2, column=2, value=Keyword)
    wsTableRows = []


    #Writing
    for building in Buildings:
        for earthquake in Earthquakes:
            SubjectCode = [building, earthquake]

            BuildingRows = rowContains_List(worksheet, building)
            EarthquakeRows = rowContains_List(worksheet, earthquake)

            SelectedTable_Rows= List_Intersection(BuildingRows, EarthquakeRows)

            # Preferred to use single Selected Cols for single sheet i.e Group_A
            SelectedCols = colContains_List(worksheet, Keyword)
            IdentifierCol = colContains_List(worksheet, Identifiers[0])[0]



            #Selected rows list (Contains the value with rowStart and rowEnd)
            SelectedRows = []
            for row in SelectedTable_Rows:
                RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(worksheet, row)
                SelectedRows.append([RowStart, RowEnd])



            #Writer
            CurrentTableRow = LiveRow
            wsTableRows.append(CurrentTableRow)
            LiveRow, XValues = GroupWriter(worksheet, ws, LiveRow, SubjectCode, IdentifierCol, SelectedRows, SelectedCols, skip_TitleRows, CustomCol = CustomColumn)




    #Make sure to save file after calling GroupWriter
    workbook.save(OutputPath)

    # Table Info
    TableInfo = []
    for row in wsTableRows:
        RowStart, RowEnd, ColumnStart, ColumnEnd = tableExtent(ws, row)
        TableInfo.append([RowStart, RowEnd, ColumnStart, ColumnEnd])

    #Correction
    if Keyword.split(" ")[0] == "Drift":
        CorrectionRow = 4
        SurrRefRow = 1
        for info in TableInfo:
            RowStart, RowEnd, ColumnStart, ColumnEnd = info[0], info[1], info[2], info[3]

            ValuesList = [[], [RowStart + skip_TitleRows, RowEnd]]
            for i in range(ColumnStart + 1, ColumnEnd):
                ValuesList[0].append(i)
            CorrType = "Average"


            Correction(ws, ValuesList, CorrectionRow,CorrType,  SurrRefRow=SurrRefRow)



    # Plotter
    for info in TableInfo:
        RowStart, RowEnd, ColumnStart, ColumnEnd = info[0], info[1], info[2], info[3]
        # YValues =  [Column ][Row Extents]
        YValues = [[ColumnStart], [RowStart + skip_TitleRows, RowEnd]]
        XValues = [[], [RowStart + skip_TitleRows, RowEnd]]
        for i in range(ColumnStart + 1, ColumnEnd):
            XValues[0].append(i)

        Anchor = f"{get_column_letter(ColumnEnd + 2)}{RowStart}"

        key = Keyword.split()[0]
        direction = Keyword.split()[1]
        BuildingName = ws.cell(row=RowStart, column=1).value
        EQName = ws.cell(row=RowStart, column=2).value
        CTitle  = f"{key} of {BuildingName} building for {EQName} earthquake along {direction}"

        if Keyword.split()[0] == "Displacement":
            XLabel = Keyword.split()[0] + " (m)"

        elif Keyword.split()[0] == "Drift":
            XLabel = Keyword.split()[0]
        else:
            XLabel = Keyword

        Charts(ws, YValues, XValues, XLabel=XLabel, YLabel=CustomColumn[0],SeriesLabel = Soils, AnchourCell=Anchor, Title=CTitle)

    workbook.save(OutputPath)





folderpath = open_folder_dialog()
subfolders = list_subfolders(folderpath)
OutputName =  "ResultAssembly"
Sheetname = "ResultAssembly"
FileName = "Result.txt"
OutputPath = os.path.join(folderpath, f'{OutputName}.xlsx')

#Writing for the Result Assembly file
# ResultFiles = filePaths(subfolders, FileName)
# ExcelWriter(ResultFiles, OutputPath, Sheetname)


Buildings = [ "L1", "L2", "L3", "L4", "R", "S"]
Earthquakes = ["Gorkha", "Northridge", "San Fernando"]
Soils = ["Fixed", "Hard", "Medium",  "Soft"]
ResultsKey = ["Drift X",'Drift Y','Displacement X','Displacement Y','Reaction Force X','Reaction Force Y','Reaction Moment X','Reaction Moment Y','Rocking Angle X','Rocking Angle Y','Rocking Angle Z','Max Uplift','Max Uplift Point','Max Settlement','Max Settlement Point']


#Customized Arranger Chart List
SeriesColour = ["D30000", "0018F9", "3BB143", "FCE205", "FC0FC0", "2B1700", "000080", "FF00FF", "4CBB17", "C21807", "0080FE"]
Markers = [ "x", "triangle", "square",  "circle",  "plus", "diamond", "star", "dash", "dot",  "picture",  "auto"]
Dashes = ["solid", "sysDash", "dash", "dot", "sysDot", "lgDashDot",  "dashDot", "lgDashDotDot", "sysDashDotDot", "lgDash", "sysDashDot"]


# for key in ResultsKey[:4]:
#     Group_A(Keyword = key)

#Returns the starting row of all tables inside of the worksheet
def Row_AllTables(worksheet):
    maxRow = worksheet.max_row

    Rows = []

    indicator = 1
    for row in range(1, 1+ maxRow):
        Value = worksheet.cell(column = 1, row = row).value

        if Value != None:
            if indicator:
                Rows.append(row)
            indicator = 0
        else:
            indicator = 1

    return Rows





# This group writes textfile which is further used for extraction of flexible based building information from STKO console.
def Group_B():
    workbook, worksheet = ExcelObj(OutputPath, Sheetname)
    TableRows = Row_AllTables(worksheet)
    writerBiasedFor = "Fixed"
    RefCols1 = colContains_List(worksheet, "Reaction Force X")
    SkipRows1 = 2

    RefCols2 = colContains_List(worksheet, "Displacement X")
    SkipRows2 = 7

    WritingFile = os.path.join(folderpath, "PseudoInfo.txt")
    file = open(WritingFile, "w+")
    for row in TableRows:
        LineValue = []
        ModelCode = ""
        for col in range(1,4):
            Value = worksheet.cell(row = row, column = col).value
            if col == 1:
                ModelCode = Value
            else:
                ModelCode  = ModelCode + "_" + Value

        Codes = ModelCode.split("_")
        if writerBiasedFor not in Codes:
            break

        LineValue.append(ModelCode)

        RefValue1 = worksheet.cell(row = row + SkipRows1, column = RefCols1[0]).value
        PseudoTime = worksheet.cell(row = row + (SkipRows1 + 1), column = RefCols1[0]).value
        RefValue2 = worksheet.cell(row = row + SkipRows2, column = RefCols2[0]).value

        LineValue.append(str(RefValue1 ))
        LineValue.append(str(PseudoTime))
        LineValue.append(str(RefValue2 ))


        file.write('\t'.join(LineValue))
        file.write('\n')


Group_B()
