import os

import numpy as np
import h5py
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
import openpyxl

def getNodeCoOrdinates(filepath, Node):
    with h5py.File(filepath, "r") as hdf:
        ls = list(hdf.keys())
        datastage = hdf.get(ls[1])
        modelinfo = datastage.get(list(datastage.keys())[0])
        nodes = modelinfo.get(list(modelinfo.keys())[0])
        nodesID =  np.array(nodes.get(list(nodes.keys())[0]))
        nodesCoOrd = np.array(nodes.get(list(nodes.keys())[1]))

        nodeAbsIndex = np.where(nodesID == Node)
        row_index = nodeAbsIndex[0][0] + 1

        CoOrdinates = nodesCoOrd[nodeAbsIndex]
        return CoOrdinates[0]
def getEleNodes(filepath, ElementID):
    with h5py.File(filepath, "r") as hdf:
        ls = list(hdf.keys())
        datastage = hdf.get(ls[1])
        modelinfo = datastage.get(list(datastage.keys())[0])
        nodes = modelinfo.get(list(modelinfo.keys())[0])
        elements = modelinfo.get(list(modelinfo.keys())[1])

        sspbrick = np.array(elements.get(list(elements.keys())[1]))
        sspbrickid = sspbrick[:, 0]
        sspbricknode = sspbrick[:, 1:9]

        absIndex = np.where(sspbrickid == ElementID)
        row_index = absIndex[0][0]
        Nodes = sspbricknode[absIndex][0]

        return absIndex, Nodes
def getMidPt(filepath, ElementID):
    absIndex, Nodes = getEleNodes(filepath, ElementID)
    CoOrdinates = [[], [], []]
    for node in Nodes:
        CoOrd = getNodeCoOrdinates(filepath, node)
        X = CoOrd[0]
        Y = CoOrd[1]
        Z = CoOrd[2]
        CoOrdinates[0].append(X)
        CoOrdinates[1].append(Y)
        CoOrdinates[2].append(Z)

    MidPoint = [sum(CoOrdinates[0]) / len(CoOrdinates[0]), sum(CoOrdinates[1]) / len(CoOrdinates[1]),
                sum(CoOrdinates[2]) / len(CoOrdinates[2])]
    return MidPoint
def get_EleIds(filepath, key):
    with h5py.File(filepath, "r") as hdf:
        datastage = hdf.get('MODEL_STAGE[2]')
        results = datastage.get('RESULTS')
        elements = results.get('ON_ELEMENTS')
        stress = elements.get(key)
        stressSoil = stress.get("121-SSPbrick[400:0:1]")

        # < KeysViewHDF5['META', 'ID', 'DATA'] >
        ids = np.array(stressSoil.get("ID"))[:,0]
        return ids

def get_Elerow(key, element_id):
    with h5py.File(filepath, "r") as hdf:
        datastage = hdf.get('MODEL_STAGE[2]')
        results = datastage.get('RESULTS')
        elements = results.get('ON_ELEMENTS')
        stress = elements.get(key)
        stressSoil = stress.get("121-SSPbrick[400:0:1]")

        # < KeysViewHDF5['META', 'ID', 'DATA'] >
        ids = np.array(stressSoil.get("ID"))[:, 0]
        absrow = np.where(ids == element_id)[0][0]

        return absrow


def FactorReturn(a , b):
    Factor = abs(a - b)
    if Factor == 0:
        Factor = 1
    return  Factor


def ExcelObj(OutputPath, Sheetname):
    # Create a new Excel workbook
    if os.path.exists(OutputPath):
        workbook = openpyxl.load_workbook(OutputPath)

        if Sheetname in workbook.sheetnames:
            worksheet = workbook[Sheetname]
        else:
            worksheet = workbook.create_sheet(title=Sheetname)

        return workbook, worksheet

    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title=Sheetname)
        return workbook, worksheet



def open_folder_dialog():
    root = tk.Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(
        title="Select the root folder containing the results"
    )

    return folder_path


def list_subfolders(root_folder):
    subfolders = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        for dirname in dirnames:
            subfolder_path = os.path.join(dirpath, dirname)
            subfolders.append(subfolder_path)
    return subfolders



def ExtractStressStrain(filepath):
    # Get Element IDS
    IDS = get_EleIds(filepath, "stress")

    #Get ids and their mid point co ordinates
    id_CoOrd = {}
    for id in IDS:
        id_CoOrd[id] =  getMidPt(filepath, id)

    #Get nearest node with reference type
    Factors = []

    for key, CoOrdinates in id_CoOrd.items():
        Factor = FactorReturn(CoOrdinates[0] , refcoOrd[0]) * FactorReturn(CoOrdinates[1] , refcoOrd[1]) * FactorReturn(CoOrdinates[2] , refcoOrd[2])
        Factors.append(Factor)


    Index = Factors.index(min(Factors))
    keylist = list(id_CoOrd.keys())
    Element = keylist[Index]
    print(Element)
    # Element = 8949
    # Stress
    absrow = get_Elerow("stress", Element)
    stresses = np.empty(0)
    strains = np.empty(0)
    with h5py.File(filepath, "r") as hdf:
        datastage = hdf.get('MODEL_STAGE[2]')
        results = datastage.get('RESULTS')
        elements = results.get('ON_ELEMENTS')

        stress = elements.get("stress")
        stressSoil = stress.get("121-SSPbrick[400:0:1]")
        stress_data = stressSoil.get("DATA")
        keys = list(stress_data.keys())
        for data_index in keys:
            datavalues = np.array(stress_data.get(data_index))
            value = datavalues[absrow, component]
            stresses = np.append(stresses, value)

        stress = elements.get("strain")
        stressSoil = stress.get("121-SSPbrick[400:0:1]")
        stress_data = stressSoil.get("DATA")
        keys = list(stress_data.keys())
        for data_index in keys:
            datavalues = np.array(stress_data.get(data_index))
            value = datavalues[absrow, component]
            strains = np.append(strains, value)

        return stresses, strains

refcoOrd = [12.5, 12.5, -7]  # Reference CoOrdinates in X, Y and Z format
Tolerances = [0.5, 2, 5]  # Toleraces: Starting tolerance for check, tolerance increment, and maximum tolerance
component = 5




folderpath = open_folder_dialog()
OutputPath = os.path.join(folderpath, "StressStrain.xlsx")
subfolders = list_subfolders(folderpath)
subfolders.append(folderpath)

#Getting all the MPC
MPCOFiles = []
for folder in subfolders:
    files = os.listdir((folder))

    #Get mpco type files and their sizes
    fileSizes = {}
    for file in files:
        if file.endswith(".mpco"):
            filepath = os.path.join(folder, file)
            file_size = os.path.getsize(filepath)  # Get the file size in bytes

            fileSizes[file] = file_size
    ValueList = list(fileSizes.values())
    KeyList = list(fileSizes.keys())
    try:
        MaxValIndex = ValueList.index(max(ValueList))
        reqKey = KeyList[MaxValIndex]
        firstName = reqKey.split(".")[0]
    except:
        pass

    #Get all the corresponding files
    reqFiles = [os.path.join(folder, x) for x in KeyList if x.split(".")[0] == firstName]
    if len(reqFiles) > 0:
        MPCOFiles.append(reqFiles)


# #Main theme is to find minimum factor (which represents for the higher closeness to the reference point) and
#         #the correponding file from filegroup so as to further proceed for data processing
ReqFiles = []
for filegroup in MPCOFiles:
    globalFactor = []
    for filepath in filegroup:

        # Get Element IDS
        print(filepath)
        IDS = get_EleIds(filepath, "stress")

        # Get ids and their mid point co ordinates
        id_CoOrd = {}
        for id in IDS:
            id_CoOrd[id] = getMidPt(filepath, id)

        # Get nearest node with reference type
        Factors = []

        for key, CoOrdinates in id_CoOrd.items():
            Factor = FactorReturn(CoOrdinates[0], refcoOrd[0]) * FactorReturn(CoOrdinates[1],
                                                                              refcoOrd[1]) * FactorReturn(
                CoOrdinates[2], refcoOrd[2])
            Factors.append(Factor)
        globalFactor.append(min(Factors))

    globalFactorIndex = globalFactor.index(min(globalFactor))
    ReqFiles.append(filegroup[globalFactorIndex])



# ReqFiles = ['D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Gorkha\\Hard\\L4_Hard_Gorkha.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Gorkha\\Medium\\L4_Medium_Gorkha.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Gorkha\\Soft\\L4_Soft_Gorkha.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Northridge\\Hard\\L4_Hard_Northridge.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Northridge\\Medium\\L4_Medium_Northridge.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\Northridge\\Soft\\L4_Soft_Northridge.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\San Fernando\\Hard\\L4_Hard_San Fernando.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\San Fernando\\Medium\\L4_Medium_San Fernando.part-1.mpco', 'D:/Final Analysis/Extras/Nonlinear Analysis/L4\\San Fernando\\Soft\\L4_Soft_San Fernando.part-1.mpco']


for file in ReqFiles:
    filepath = Path(file)
    Code =  ("_").join(filepath.parts[-3:-1])
    stress, strain = ExtractStressStrain(filepath)

    wb, ws  = ExcelObj(OutputPath, Code)

    for index, row in enumerate(range(1, len(stress) + 1)):
        ws.cell(row = row, column = 1, value = strain[index])
        ws.cell(row = row, column = 2, value = stress[index])

    wb.save(OutputPath)




#
# # Create a basic line plot
# plt.plot(strains, stresses)
#
# # Add labels and a title
# plt.xlabel('X-axis')
# plt.ylabel('Y-axis')
# plt.title('Simple Plot')
#
# # Show the plot
# plt.show()
#


